import datetime
import requests
from urllib.parse import urlencode
from openpyxl import load_workbook

from models import Application
from message_text import ButtonApplication


#  регистрация новах пользователей в файл
def start_registration(message):
    make_start_time = datetime.datetime.now()
    make_start_time += datetime.timedelta(hours=3)  # Перевод в Московское время
    make_start_time = make_start_time.strftime('%d.%m.%Y-%H:%M')
    with open('data/mainFiles/list_registration.txt', 'w', encoding='utf-8') as f:
        print(message.from_user.id, make_start_time, sep=';', file=f)
        print(message.from_user.id, make_start_time, sep=';')


# Скачивает файл с яндекс диска по ссылке на файл -> возвращает новый src
def download_from_yadisk(link, file_name="test1.xlsx"):
    base_url = 'https://cloud-api.yandex.net/v1/disk/public/resources/download?'

    # Получаем загрузочную ссылку
    final_url = base_url + urlencode(dict(public_key=link))
    response = requests.get(final_url)
    download_url = response.json()['href']

    # Загружаем файл и сохраняем его
    download_response = requests.get(download_url)
    src = f'data/tempFiles/{file_name}'
    with open(src, 'wb') as f:  # Здесь укажите нужный путь к файлу
        f.write(download_response.content)
    return src


# Принимает путь к эксель файлу -> возвращает список действующих компаний
def set_list_organization(src):
    wb = load_workbook(src, data_only=True)
    sheet = wb.active
    organization_list = []
    row = 2
    while True:
        cell_value = sheet[f"B{row}"].value
        if not cell_value:
            break
        if sheet[f"C{row}"].value:
            organization_list.append(cell_value)
        row += 1
    return organization_list


# Принимает путь к эксель файлу + название компании  ->  возвращает id менеджера
def set_manager_id(src, organization):
    wb = load_workbook(src, data_only=True)
    sheet = wb.active
    manager_id = None
    row = 2
    while True:
        cell_value = sheet[f"B{row}"].value
        if not cell_value:
            break
        elif cell_value == organization:
            manager_id = sheet[f"D{row}"].value
        row += 1
    return manager_id


# Принимает список заявок  ->  возвращает xlsx файл с
def fill_file(app_list: list[Application]):
    wb = load_workbook("data/mainFiles/applications.xlsx", data_only=True)
    sheet = wb.active
    row = 5
    for i in range(len(app_list)):
        application = app_list[i]
        product = application.product
        the_object = application.the_object
        manager = application.manager
        client = application.creator

        if application.unloading_method == ButtonApplication.unloading_method_list[0]:
            speed = the_object.average_speed_of_unloading_by_crane
        elif application.unloading_method == ButtonApplication.unloading_method_list[1]:
            speed = the_object.average_speed_of_unloading_by_pump
        else:
            speed = the_object.average_speed_of_unloading_by_self_watering

        sheet[f"A{row}"] = i + 1
        sheet[f"B{row}"] = application.execution_time.strftime("%H:%M")
        sheet[f"C{row}"] = application.id
        sheet[f"D{row}"] = application.is_executed
        sheet[f"E{row}"] = application.is_confirmation
        sheet[f"F{row}"] = product.abbreviation
        sheet[f"G{row}"] = application.volume_declared
        sheet[f"H{row}"] = the_object.id
        sheet[f"I{row}"] = the_object.obj_name
        sheet[f"J{row}"] = application.unloading_method
        sheet[f"K{row}"] = application.declared_unloading_speed
        sheet[f"L{row}"] = speed
        sheet[f"M{row}"] = f"{client.name}  {client.last_name}"
        sheet[f"N{row}"] = client.organization
        sheet[f"O{row}"] = f"{manager.name}  {manager.last_name}"
        row += 1
    src = f"data/userFiles/uploading_applications/app_{datetime.datetime.now().strftime('%H_%M_%S')}.xlsx"
    wb.save(src)
    return src
